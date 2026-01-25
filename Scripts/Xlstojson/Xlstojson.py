# -*- coding: utf-8 -*-
import json
import openpyxl
import os
import re
from datetime import datetime, date, time, timedelta
from typing import Any, Dict, Optional, List


HEADER_ROW = 6
DATA_START_ROW = HEADER_ROW + 1

# ------------------------
# Helpers : base
# ------------------------

def _norm_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()

def _norm_numlike(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        return ("%.15g" % v).strip()
    return str(v).strip()

def _fold(s: Any) -> str:
    import unicodedata
    t = _norm_str(s).lower()
    t = unicodedata.normalize("NFD", t)
    t = "".join(ch for ch in t if not unicodedata.combining(ch))
    t = t.replace("_", " ").replace("-", " ")
    t = " ".join(t.split())
    return t

def _to_date_str(v: Any) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, date):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, (int, float)):
        n = int(v)
        if 20000 <= n <= 60000:
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=n)).strftime("%d/%m/%Y")
        s = _norm_str(v)
    else:
        s = _norm_str(v)
    # Supprimer TOUS les guillemets (simples et doubles)
    s = s.replace('"', '').replace("'", "")
    s = s.replace("T", " ").split(" ")[0].strip()
    m = re.match(r"^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})$", s)
    if m:
        y, mo, d = m.groups()
        return f"{int(d):02d}/{int(mo):02d}/{int(y):04d}"
    m = re.match(r"^(\d{1,2})[-/.](\d{1,2})[-/.](\d{4})$", s)
    if m:
        d, mo, y = m.groups()
        return f"{int(d):02d}/{int(mo):02d}/{int(y):04d}"
    return s

def replace_none_with_empty_string(data):
    if isinstance(data, dict):
        for key in list(data.keys()):
            if data[key] is None:
                data[key] = ""
            else:
                replace_none_with_empty_string(data[key])
    elif isinstance(data, list):
        for item in data:
            replace_none_with_empty_string(item)

def convert_datetime_and_numeric(data):
    if isinstance(data, dict):
        for key, value in list(data.items()):
            if isinstance(value, datetime):
                data[key] = value.strftime("%d/%m/%Y")
            elif isinstance(value, time):
                data[key] = value.strftime("%H:%M:%S")
            elif isinstance(value, (int, float)):
                data[key] = _norm_numlike(value)
            else:
                convert_datetime_and_numeric(value)
    elif isinstance(data, list):
        for i, item in enumerate(list(data)):
            if isinstance(item, datetime):
                data[i] = item.strftime("%d/%m/%Y")
            elif isinstance(item, time):
                data[i] = item.strftime("%H:%M:%S")
            elif isinstance(item, (int, float)):
                data[i] = _norm_numlike(item)
            else:
                convert_datetime_and_numeric(item)

# ------------------------
# NORMALISATIONS — TES RÈGLES
# ------------------------

def normalize_prenom(s: str) -> str:
    if not s:
        return ""
    parts = re.split(r"([-'])", s)  # garde tirets/apostrophes
    return "".join(p.capitalize() if p.isalpha() else p for p in parts)

def normalize_pseudo_list(ayants):
    """Si un pseudo 100% MAJ → tous en MAJ, sinon Title Case."""
    pseudos = [
        ad["Identite"].get("Pseudonyme", "")
        for ad in ayants
        if ad["Identite"].get("Type", "").lower() != "moral"
    ]
    if not pseudos:
        return False

    has_full_upper = any(p and p.isupper() for p in pseudos)

    for ad in ayants:
        ident = ad["Identite"]
        if ident.get("Type", "").lower() == "moral":
            continue

        p = ident.get("Pseudonyme", "")
        if not p:
            continue

        if has_full_upper:
            ident["Pseudonyme"] = p.upper()
        else:
            ident["Pseudonyme"] = normalize_prenom(p)

def post_process(data):
    ayants = data.get("AyantsDroit", [])

    # 1. normalisation globale des pseudonymes
    normalize_pseudo_list(ayants)

    for ad in ayants:
        ident = ad.get("Identite", {})
        adr = ad.get("Adresse", {})

        # 2. Pays → upper()
        if "Pays" in adr and adr["Pays"]:
            adr["Pays"] = adr["Pays"].upper()

        # 3. Type Physique
        if ident.get("Type", "").lower() == "physique":
            # Nom → upper
            if ident.get("Nom"):
                ident["Nom"] = ident["Nom"].upper()
            # Prénom → Title Case + tirets/apostrophes
            if ident.get("Prenom"):
                ident["Prenom"] = normalize_prenom(ident["Prenom"])

        # 4. Type Moral → prénom/nom représentant
        if ident.get("Type", "").lower() == "moral":
            if ident.get("PrenomRepresentant"):
                ident["PrenomRepresentant"] = normalize_prenom(
                    ident["PrenomRepresentant"]
                )
            if ident.get("NomRepresentant"):
                ident["NomRepresentant"] = ident["NomRepresentant"].upper()

# ------------------------
# Mapping en-têtes
# ------------------------

EXPECTED = {
    "role": ["role"],
    "designation": ["designation"],
    "lettrage": ["lettrage"],
    "coad_ipi": ["coad/ipi", "coad", "ipi", "coad ipi"],
    "ph": ["ph"],
    "managelic": ["managelic"],
    "managesub": ["managesub"],
    "societegestion": ["societegestion", "société gestion", "societe gestion"],
    "prenom": ["prenom", "prénom"],
    "nom": ["nom"],
    "type": ["type"],
    "pseudonyme": ["pseudonyme", "pseudo", "alias"],
    "genre": ["genre"],
    "nele": ["nele", "néle", "né le", "ne le"],
    "nea": ["nea", "néa", "né à", "ne a", "né a"],
    "de": ["de"],
    "dr": ["dr"],
    "numvoie": ["numvoie", "num voie", "numéro de voie", "numero de voie", "n° voie"],
    "typevoie": ["typevoie", "type voie", "type de voie"],
    "nomvoie": ["nomvoie", "nom voie", "nom de la voie", "voie", "rue", "avenue"],
    "cp": ["cp", "code postal"],
    "ville": ["ville", "commune", "localité", "localite", "city"],
    "pays": ["pays", "country"],
    "mail": ["mail", "email", "e-mail", "courriel"],
    "tel": ["tel", "tél", "telephone", "téléphone"],
    "formejuridique": ["formejuridique", "forme juridique"],
    "capital": ["capital"],
    "rcs": ["rcs"],
    "siren": ["siren"],
    "genrerepresentant": ["genrerepresentant", "genrereprestant", "genre representant", "genre représentant"],
    "prenomrepresentant": ["prénomrepresentant", "prenomrepresentant",
                           "prénom representant", "prenom representant"],
    "nomrepresentant": ["nomrepresentant", "nom representant"],
    "fonctionrepresentant": ["fonctionrepresentant", "fonction representant"],
}

def build_header_map(sheet) -> Dict[str, Optional[int]]:
    max_cols = sheet.max_column
    folded_to_col = {}
    for col_idx in range(1, max_cols + 1):
        folded_to_col[_fold(sheet.cell(row=HEADER_ROW, column=col_idx).value)] = col_idx

    def find_by_synonyms(syns):
        for s in syns:
            key = _fold(s)
            if key in folded_to_col:
                return folded_to_col[key]
        return None

    return {k: find_by_synonyms(v) for k, v in EXPECTED.items()}

# ------------------------
# Recherche étiquettes
# ------------------------

def _find_label_value(ws, label_text: str) -> str:
    target = _fold(label_text)
    max_row_scan = max(1, min(ws.max_row, (HEADER_ROW - 1) if HEADER_ROW > 1 else 20))
    max_col = ws.max_column
    for r in range(1, max_row_scan + 1):
        for c in range(1, max_col + 1):
            if _fold(ws.cell(r, c).value) == target:
                if c + 1 <= max_col:
                    right = _norm_str(ws.cell(r, c + 1).value)
                    if right:
                        return right
                if r + 1 <= ws.max_row:
                    below = _norm_str(ws.cell(r + 1, c).value)
                    if below:
                        return below
                return ""
    return ""

# ------------------------
# Fallback choix éditeur
# ------------------------

def _pick_editors_fallback(ayants: List[Dict[str, Any]]) -> str:
    editors = {}
    order = []
    for ad in ayants:
        bdo = ad.get("BDO", {}) or {}
        ident = ad.get("Identite", {}) or {}
        if (bdo.get("Role") or "").upper() != "E":
            continue
        # Pour Moral: Designation, Pour Physique: construire depuis Nom/Prenom
        if ident.get("Type", "").lower() == "moral":
            name = (ident.get("Designation") or "").strip()
        else:
            nom = (ident.get("Nom") or "").strip()
            prenom = (ident.get("Prenom") or "").strip()
            name = f"{nom} {prenom}".strip()
        if not name:
            continue
        if name not in editors:
            editors[name] = {
                "is_sacem": ((ident.get("SocieteGestion") or "").strip().upper() == "SACEM"),
                "ph_total": 0.0,
                "manages": 0,
                "first": len(order),
            }
            order.append(name)
        try:
            editors[name]["ph_total"] += float(str(bdo.get("PH", "0")).replace(",", "."))
        except Exception:
            pass

    for ad in ayants:
        bdo = ad.get("BDO", {}) or {}
        if (bdo.get("Role") or "").upper() != "E":
            continue
        mgr = (bdo.get("Managelic") or "").strip()
        if mgr and mgr in editors:
            editors[mgr]["manages"] += 1

    if not editors:
        return ""

    def score(name):
        e = editors[name]
        return (1 if e["is_sacem"] else 0, e["manages"], e["ph_total"], -e["first"])

    best = sorted(editors.keys(), key=score, reverse=True)[0]
    return best

# ------------------------
# Extraction principale
# ------------------------

def extract_data_from_xlsx(file_path: str) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sh = wb.active

    headers = build_header_map(sh)

    data = {
        "Titre": _norm_str(sh['B1'].value),
        "SousTitre": _norm_str(sh['B2'].value),
        "Interprete": _norm_str(sh['B3'].value),
        "Duree": _norm_str(sh['B4'].value),
        "Genre": _norm_str(sh['B5'].value),

        "Date": _to_date_str(sh['D1'].value),
        "ISWC": _norm_str(sh['D2'].value),
        "Lieu": _norm_str(sh['D3'].value),
        "Territoire": _norm_str(sh['D4'].value),
        "Arrangement": _norm_str(sh['D5'].value),

        "Inegalitaire": _norm_str(sh['F1'].value),
        "Commentaire": _norm_str(sh['F2'].value),
        "Faita": _norm_str(sh['F3'].value),
        "Faitle": _to_date_str(sh['F4'].value),

        "Declaration": _norm_str(sh['I1'].value),
        "Format": _norm_str(sh['I2'].value),

        "AyantsDroit": []
    }

    def gv(row: int, key: str, kind: str = "text") -> str:
        col = headers.get(key)
        val = sh.cell(row=row, column=col).value if col else ""
        return _norm_numlike(val) if kind == "num" else _norm_str(val)

    max_row = sh.max_row
    for row in range(DATA_START_ROW, max_row + 1):
        role = gv(row, "role")
        designation = gv(row, "designation")
        if not role and not designation:
            continue

        ident_type = gv(row, "type")

        # Structure différente selon Type
        if _fold(ident_type) == "moral":
            # === TYPE MORAL : avec Designation ===
            ident: Dict[str, Any] = {
                "Designation": designation,
                "Type": "Moral",
                "SocieteGestion": gv(row, "societegestion"),
                "FormeJuridique": gv(row, "formejuridique"),
                "Capital": gv(row, "capital"),
                "RCS": gv(row, "rcs"),
                "Siren": gv(row, "siren"),
                "GenreRepresentant": gv(row, "genrerepresentant"),
                "PrenomRepresentant": gv(row, "prenomrepresentant"),
                "NomRepresentant": gv(row, "nomrepresentant"),
                "FonctionRepresentant": gv(row, "fonctionrepresentant"),
            }
        else:
            # === TYPE PHYSIQUE : SANS Designation ===
            ident: Dict[str, Any] = {
                "Type": "Physique",
                "SocieteGestion": gv(row, "societegestion"),
                "Pseudonyme": gv(row, "pseudonyme"),
                "Nom": gv(row, "nom"),
                "Prenom": gv(row, "prenom"),
                "Genre": gv(row, "genre"),
                "Nele": _to_date_str(gv(row, "nele")),
                "Nea": gv(row, "nea"),
            }

        ayant = {
            "Identite": ident,
            "BDO": {
                "Role": role,
                "Lettrage": gv(row, "lettrage"),
                "COAD/IPI": gv(row, "coad_ipi"),
                "PH": gv(row, "ph"),
                "Managelic": gv(row, "managelic"),
                "Managesub": gv(row, "managesub"),
                "DE": gv(row, "de"),
                "DR": gv(row, "dr"),
            },
            "Adresse": {
                "NumVoie": gv(row, "numvoie", kind="num"),
                "TypeVoie": gv(row, "typevoie"),
                "NomVoie": gv(row, "nomvoie"),
                "CP": gv(row, "cp", kind="num"),
                "Ville": gv(row, "ville"),
                "Pays": gv(row, "pays"),
            },
            "Contact": {
                "Mail": gv(row, "mail"),
                "Tel": gv(row, "tel"),
            }
        }
        data["AyantsDroit"].append(ayant)

    placeholders = {"", "declaration", "format"}
    if _fold(data.get("Declaration")) in placeholders:
        found = _find_label_value(sh, "Declaration")
        if found:
            data["Declaration"] = found
    if _fold(data.get("Format")) in placeholders:
        found = _find_label_value(sh, "Format")
        if found:
            data["Format"] = found

    if not _norm_str(data.get("Declaration")):
        data["Declaration"] = _pick_editors_fallback(data["AyantsDroit"])
    if not _norm_str(data.get("Format")):
        data["Format"] = _pick_editors_fallback(data["AyantsDroit"])

    if isinstance(data.get("Duree"), str):
        data["Duree"] = data["Duree"].strip().strip('"').strip("'")

    replace_none_with_empty_string(data)
    convert_datetime_and_numeric(data)

    # → APPLICATION DE TES RÈGLES
    post_process(data)

    return data

# ------------------------
# Batch dossier
# ------------------------

def process_files_in_directory(directory_path: str, output_directory: str):
    for file_name in os.listdir(directory_path):
        if not file_name.endswith('.xlsx'):
            continue
        if file_name.startswith("~$"):
            continue

        file_path = os.path.join(directory_path, file_name)
        data = extract_data_from_xlsx(file_path)
        titre = (data.get("Titre") or "").replace("/", "_").replace("\\", "_")
        interpr = (data.get("Interprete") or "").replace("/", "_").replace("\\", "_")
        base = f"CONF_{titre}_{interpr}".strip("_")
        output_file_name = f"{base}.json" if base else (os.path.splitext(file_name)[0] + "_.json")
        output_file_path = os.path.join(output_directory, output_file_name)
        with open(output_file_path, 'w', encoding='utf-8') as json_file:
            json.dump(data, json_file, indent=4, ensure_ascii=False)
        print(f"Processed: {file_name} -> {output_file_name}")

# ------------------------
# Entrée principale
# ------------------------

if __name__ == "__main__":
    import sys
    
    # Utilisation : python Xlstojson.py [dossier_xlsx] [dossier_json]
    # Par défaut : xlsx -> json
    
    if len(sys.argv) >= 3:
        input_directory = sys.argv[1]
        output_directory = sys.argv[2]
    elif len(sys.argv) == 2:
        input_directory = sys.argv[1]
        output_directory = 'json'
    else:
        input_directory = 'xlsx'
        output_directory = 'json'
    
    os.makedirs(output_directory, exist_ok=True)
    process_files_in_directory(input_directory, output_directory)
    print(f"\nTerminé ! Fichiers JSON générés dans : {output_directory}")

